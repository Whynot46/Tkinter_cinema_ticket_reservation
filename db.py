from openpyxl import load_workbook


class Data_base():
    wb_films= load_workbook('./db/films.xlsx')
    films_sheet = wb_films['Фильмы']
    auditorium_sheet = wb_films['Залы']
    schedule_sheet = wb_films['Расписание_фильмов']
    auditorium_1 = wb_films['Зал_1']
    auditorium_2 = wb_films['Зал_2']
    auditorium_3 = wb_films['Зал_3']
    auditorium_4 = wb_films['Зал_4']
    auditorium_5 = wb_films['Зал_5']

    wb_users= load_workbook('./db/users.xlsx')
    users_sheet = wb_users['users']

    #Сохранение базы данных
    def save_data_base(self):
        self.wb_films.save('./db/films.xlsx')
        self.wb_users.save('./db/users.xlsx')
 