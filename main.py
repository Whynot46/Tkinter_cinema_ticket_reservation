import os
from PIL import Image
import customtkinter as CTk
from tkinter import messagebox
from openpyxl import load_workbook


base_padding = {'padx': 10, 'pady': 8}
header_padding = {'padx': 10, 'pady': 12}

CTk.set_appearance_mode("System")  # Modes: system (default), light, dark
CTk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green

#Локальные данные пользователя
class User():
    def __init__(self):
        self.name = 'Guest'
        self.login = 'None'
        self.password = ''
    #Обновление локальных пользовательских данных
    def change_user_data(self, login):
        line = 1
        while line!=Data_base().users_sheet[f'I2'].value+2:
            if Data_base().users_sheet[f'C{line}'].value == login:
                self.name = Data_base().users_sheet[f'B{line}'].value
                self.login = Data_base().users_sheet[f'C{line}'].value
                self.password = Data_base().users_sheet[f'D{line}'].value
                break
            line+=1

#Окно авторизации
class Login_window():
    def __init__(self):
        self.window = CTk.CTk()
        self.window.title('Авторизация')
        self.window.geometry('350x400')
        self.window.resizable(False, False)

        self.cinema_icon_img = CTk.CTkImage(light_image=Image.open('./img/icons/login_icon.png'),
                                  dark_image=Image.open('./img/icons/login_icon.png'),
                                  size=(128, 128))
        self.user_login()


    def user_login(self):
        self.cinema_icon = CTk.CTkLabel(self.window, text="", image=self.cinema_icon_img)
        self.cinema_icon.pack()

        self.authorization_label = CTk.CTkLabel(self.window, text='Авторизация', **header_padding)
        self.authorization_label.pack()

        # метка для поля ввода имени
        self.login_label = CTk.CTkLabel(self.window, text='Логин', **base_padding)
        self.login_label.pack()

        # поле ввода имени
        self.login_entry = CTk.CTkEntry(self.window)
        self.login_entry.pack()
        self.login_entry.focus()

        # метка для поля ввода пароля
        self.password_label = CTk.CTkLabel(self.window, text='Пароль', **base_padding)
        self.password_label.pack()

        # поле ввода пароля
        self.password_entry = CTk.CTkEntry(self.window)
        self.password_entry.pack()

        # кнопка отправки формы
        self.send_btn = CTk.CTkButton(self.window, text='Войти', command=self.login)
        self.send_btn.pack(**base_padding)

        self.registration_button = CTk.CTkButton(self.window, text='Регистрация', command=self.to_registration)
        self.registration_button.pack(**base_padding)

        self.window.protocol("WM_DELETE_WINDOW", self.on_exit)

        self.window.mainloop()


    def login(self):
        login = self.login_entry.get()
        password = self.password_entry.get()
        all_ok = False
        line = 1

        if login=='' or password=='':
            messagebox.showerror('Ошибка', 'Заполните все поля')
        else:
            while line!=Data_base().users_sheet[f'I2'].value+2:
                if (Data_base().users_sheet[f'C{line}'].value == login) and (Data_base().users_sheet[f'D{line}'].value == password):
                    all_ok = True
                    User.change_user_data(User, login)
                    self.window.destroy()
                    Main_window()
                line+=1
            if not all_ok: messagebox.showerror('Ошибка', 'Неверный логин или пароль')


    def to_registration(self):
        Registration_window()


    def on_exit(self):
        if messagebox.askokcancel("Выход", "Вы действительно хотите выйти?"):
            self.window.destroy()

#Окно регистрации
class Registration_window():
    def __init__(self):
        self.window = CTk.CTkToplevel()
        self.window.title("Регистрация")
        self.window.geometry("300x465")
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.on_exit)
        self.user_registration()

    def user_registration(self):

        self.authorization_icon_img = CTk.CTkImage(light_image=Image.open('./img/icons/new_user_icon.png'),
                                  dark_image=Image.open('./img/icons/new_user_icon.png'),
                                  size=(128, 128))
        
        self.authorization_icon = CTk.CTkLabel(self.window, text="", image=self.authorization_icon_img)
        self.authorization_icon.pack()

        self.authorization_label = CTk.CTkLabel(self.window, text='Регистрация', **header_padding)
        # помещаем виджет в окно по принципу один виджет под другим
        self.authorization_label.pack()

        # метка для поля ввода имени
        self.name_label = CTk.CTkLabel(self.window, text='ФИО', **base_padding)
        self.name_label.pack()

        # поле ввода имени
        self.name_entry = CTk.CTkEntry(self.window)
        self.name_entry.pack()
        self.name_entry.focus()

        self.login_label = CTk.CTkLabel(self.window, text='Логин', **base_padding)
        self.login_label.pack()

        # поле ввода имени
        self.login_entry = CTk.CTkEntry(self.window)
        self.login_entry.pack()
        self.login_entry.focus()

        # метка для поля ввода пароля
        self.password_label = CTk.CTkLabel(self.window, text='Пароль', **base_padding)
        self.password_label.pack()

        # поле ввода пароля
        self.password_entry = CTk.CTkEntry(self.window)
        self.password_entry.pack()

        # метка для поля ввода пароля
        self.repeat_password_label = CTk.CTkLabel(self.window, text='Повторите пароль', **base_padding)
        self.repeat_password_label.pack()

        # поле ввода пароля
        self.repeat_password_entry = CTk.CTkEntry(self.window)
        self.repeat_password_entry.pack()

        # кнопка отправки формы
        self.send_btn = CTk.CTkButton(self.window, text='Принять', command=self.registration)
        self.send_btn.pack(**base_padding)


    def registration(self):
        name = self.name_entry.get()
        login = self.login_entry.get()
        password = self.password_entry.get()
        repeat_password = self.repeat_password_entry.get()

        if login=='' or password=='' or repeat_password=='':
            messagebox.showerror('Ошибка', 'Заполните все поля')
        elif password != repeat_password:
            messagebox.showerror('Ошибка', 'Пароли не совпадают')
        else:
            current_line = Data_base.users_sheet[f'I2'].value+2
            Data_base.users_sheet[f'A{current_line}'] = Data_base().users_sheet[f'I2'].value+1
            Data_base.users_sheet[f'B{current_line}'] = name
            Data_base.users_sheet[f'C{current_line}'] = login
            Data_base.users_sheet[f'D{current_line}'] = password
            Data_base.users_sheet[f'I2'] = current_line - 1
            User.change_user_data(User, login)
            messagebox.showinfo('Успешно', f'Логин: {name}\nПароль: {password}')
            Data_base.save_data_base(Data_base)
            self.window.destroy()



    def on_exit(self):
        if messagebox.askokcancel("Прервать регистрацию", "Вы действительно хотите прервать регистрацию?"):
            self.window.destroy()
            Login_window()

#Главное окно
class Main_window():
    def __init__(self):
        self.window = CTk.CTk()
        self.window.title('Резервирование билетов в кинотеатр')
        #self.window.state('zoomed') #Под Windows, если не раскрывается в полный экран, заккоментируйте следующую строчку
        self.window.attributes('-zoomed', True)
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(1, weight=1)
        self.window.protocol("WM_DELETE_WINDOW", self.on_exit)

        self.top_bar_frame = CTk.CTkFrame(self.window, corner_radius=0)
        self.top_bar_frame.grid(row=0, column=0, sticky="nsew")
        self.top_bar_frame.grid_rowconfigure(4, weight=1)

        self.poster_btn = CTk.CTkButton(self.top_bar_frame, corner_radius=0, height=100, border_spacing=10, text='Афиша',fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w", command=self.open_poster_window)
        self.poster_btn.grid(row=0, column=0, sticky="ew")

        self.cinema_btn = CTk.CTkButton(self.top_bar_frame, corner_radius=0, height=100, border_spacing=10, text='Кинотеатр',fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w", command=self.open_cinema_window)
        self.cinema_btn.grid(row=1, column=0, sticky="ew")

        self.account_btn = CTk.CTkButton(self.top_bar_frame, corner_radius=0, height=100, border_spacing=10, text='Личный кабинет',fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w", command=self.open_account_window)
        self.account_btn.grid(row=2, column=0, sticky="ew")

        if (User.name != 'Guest') and (User.login != 'None'):
            self.account_btn = CTk.CTkButton(self.top_bar_frame, text='Выйти', command=self.logout)
            self.account_btn.grid(row=9, column=0, sticky="ew", pady=5)
        else:
            self.account_btn = CTk.CTkButton(self.top_bar_frame, text='Войти', command=self.logout)
            self.account_btn.grid(row=8, column=0, sticky="ew", pady=5)
            self.account_btn = CTk.CTkButton(self.top_bar_frame, text='Зарегистрируйтесь', command=self.registration)
            self.account_btn.grid(row=9, column=0, sticky="ew", pady=5)


        self.appearance_mode_menu = CTk.CTkOptionMenu(self.top_bar_frame, values=['Dark', 'Light', 'System'], command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=10, column=0, padx=20, pady=20, sticky="s")

        self.current_frame = CTk.CTkFrame(self.window, corner_radius=0, fg_color="transparent")
        self.current_frame.grid(row=0, column=1)

        self.open_cinema_window()


    def open_poster_window(self):

        self.current_frame.destroy()

        self.current_frame = CTk.CTkFrame(self.window, corner_radius=0, fg_color="transparent")
        self.current_frame.grid(row=0, column=1)

        self.film_frame_1 = CTk.CTkFrame(self.current_frame, fg_color="transparent")
        self.film_frame_1.grid(row=0, column=0, sticky="w", pady=5)

        film_icon_img_1 = Film.current_film_icon(Film, 4)
        film_icon_1 = CTk.CTkLabel(self.film_frame_1, text="", justify = 'left', image=film_icon_img_1)
        film_icon_1.grid(row=0, column=0,rowspan=4, sticky="w", padx=1, pady=1)
        film_name_1 = CTk.CTkLabel(self.film_frame_1, height=10, justify = 'left', text=Film.current_film_name(Film, 4))
        film_name_1.grid(row=1, column=1, padx=1, pady=1, sticky='nw')
        film_description_1 = CTk.CTkLabel(self.film_frame_1, text=Film.current_film_description(Film, 4))
        film_description_1.grid(row=2, column=1, padx=1, pady=1, sticky='nw')
        buy_ticket_btn_1 = CTk.CTkButton(self.film_frame_1, text='Купить билет', command=lambda: self.buy_ticket(4))
        buy_ticket_btn_1.grid(row=3, column=1, sticky="s")


        self.film_frame_2 = CTk.CTkFrame(self.current_frame, fg_color="transparent")
        self.film_frame_2.grid(row=1, column=0, sticky="w", pady=5)
        #self.film_frame_2.grid_columnconfigure(0, weight=1)

        film_icon_img_2 = Film.current_film_icon(Film, 2)
        film_icon_2 = CTk.CTkLabel(self.film_frame_2, text="", image=film_icon_img_2)
        film_icon_2.grid(row=0, column=0, rowspan=4, sticky="w", padx=1, pady=1)
        film_name_2 = CTk.CTkLabel(self.film_frame_2, height=10, justify = 'left', text=Film.current_film_name(Film, 2))
        film_name_2.grid(row=0, column=1, padx=1, pady=1, sticky='nw')
        film_description_2 = CTk.CTkLabel(self.film_frame_2, text=Film.current_film_description(Film, 2))
        film_description_2.grid(row=1, column=1, padx=1, pady=1, sticky='nw')
        buy_ticket_btn_2 = CTk.CTkButton(self.film_frame_2, text='Купить билет', command=lambda: self.buy_ticket(2))
        buy_ticket_btn_2.grid(row=2, column=1, sticky="s")


        self.film_frame_3 = CTk.CTkFrame(self.current_frame, fg_color="transparent")
        self.film_frame_3.grid(row=2, column=0, sticky="w", pady=5)
        #self.film_frame_3.grid_columnconfigure(0, weight=1)

        film_icon_img_3 = Film.current_film_icon(Film, 3)
        film_icon_3 = CTk.CTkLabel(self.film_frame_3, text="", image=film_icon_img_3)
        film_icon_3.grid(row=0, column=0, rowspan=4, sticky="w", padx=1, pady=1)
        film_name_3 = CTk.CTkLabel(self.film_frame_3, height=10, justify = 'left', text=Film.current_film_name(Film, 3))
        film_name_3.grid(row=0, column=1, padx=1, pady=1, sticky='nw')
        film_description_2 = CTk.CTkLabel(self.film_frame_3, text=Film.current_film_description(Film, 3))
        film_description_2.grid(row=1, column=1, padx=1, pady=1, sticky='nw')
        buy_ticket_btn_3 = CTk.CTkButton(self.film_frame_3, text='Купить билет', command=lambda: self.buy_ticket(3))
        buy_ticket_btn_3.grid(row=2, column=1, sticky="s")

    def open_cinema_window(self): 

        self.current_frame.destroy()

        self.current_frame = CTk.CTkFrame(self.window, corner_radius=0, fg_color="transparent")
        self.current_frame.grid(row=0, column=1)

        cinema_icon_img = CTk.CTkImage(light_image=Image.open('img/icons/cinema_icon.png'),
                                  dark_image=Image.open('img/icons/cinema_icon.png'),
                                  size=(256, 256))
        cinema_icon = CTk.CTkLabel(self.current_frame, text="", image=cinema_icon_img)
        cinema_icon.pack()
        cinema_name = CTk.CTkLabel(self.current_frame, text='Кинотеатр "Художественный фильм"')
        cinema_name.pack()

    def open_account_window(self):

        self.current_frame.destroy()

        self.current_frame = CTk.CTkFrame(self.window, corner_radius=0, fg_color="transparent")
        self.current_frame.grid(row=0, column=1)

        user_icon_img = CTk.CTkImage(light_image=Image.open('img/icons/user_icon.png'),
                                  dark_image=Image.open('img/icons/user_icon.png'),
                                  size=(256, 256))
        user_icon = CTk.CTkLabel(self.current_frame, text="", image=user_icon_img)
        user_icon.pack()
        user_name = CTk.CTkLabel(self.current_frame, text=User.login)
        user_name.pack()

    def change_appearance_mode_event(self, new_appearance_mode):
        CTk.set_appearance_mode(new_appearance_mode)

    def buy_ticket(self, film_id): 
        Buy_ticket(film_id)

    def logout(self):
        self.window.destroy()
        Login_window()

    def registration(self):
        self.window.destroy()
        Registration_window()

    def on_exit(self):
        if messagebox.askokcancel("Выход", "Вы действительно хотите выйти?"):
            self.window.destroy()

#База данных
class Data_base():
    wb_films= load_workbook('./db/films.xlsx')
    films_sheet = wb_films['Фильмы']
    auditorium_sheet = wb_films['Залы']
    schedule_sheet = wb_films['Расписание_фильмов']

    wb_users= load_workbook('./db/users.xlsx')
    users_sheet = wb_users['users']

    def save_data_base(self):
        self.wb_films.save('./db/films.xlsx')
        self.wb_users.save('./db/users.xlsx')
 

class Buy_ticket():
    def __init__(self, film_id):
        self.window = CTk.CTkToplevel()
        self.window.title("Бронирование")
        self.window.geometry('350x370')
        self.window.resizable(False, False)
        self.create_buy_ticket_form(film_id)

    def create_buy_ticket_form(self, film_id):

        film_name = CTk.CTkLabel(self.window, text='<Название_фильма>  6+', **base_padding)
        film_name.pack()

        time_label = CTk.CTkLabel(self.window, text='Время', **base_padding)
        time_label.pack()
        time_arr = ['9:00', '11:30', '14:00']
        time = CTk.CTkOptionMenu(self.window, values=time_arr)
        time.pack()

        row_label = CTk.CTkLabel(self.window, text='Ряд', **base_padding)
        row_label.pack()
        row_arr = list(map(str, range(1, 5)))
        row = CTk.CTkOptionMenu(self.window, values=row_arr)
        row.pack()

        place_label = CTk.CTkLabel(self.window, text='Место', **base_padding)
        place_label.pack()
        place_arr = list(map(str, range(1, 5)))
        place = CTk.CTkOptionMenu(self.window, values=place_arr)
        place.pack()

        username_label = CTk.CTkLabel(self.window, text='ФИО зрителя', **base_padding)
        username_label.pack()
        username_entry = CTk.CTkEntry(self.window)
        username_entry.pack()

        price = CTk.CTkLabel(self.window, text='Цена билета: <цена>\n(оплата при входе в зрительный зал)', **base_padding)
        price.pack()

        buy_btn = CTk.CTkButton(self.window, text='Купить')
        buy_btn.pack()
    

        




#Фильмы
class Film():
    def __init__(self):

        self.icon = CTk.CTkImage(light_image=Image.open('img/icons/film_icon.png'),
                                  dark_image=Image.open('img/icons/film_icon.png'),
                                  size=(128, 128))
        self.name = 'Name'
        self.description = 'Description'

    def current_line(self, id):
        line = 1
        while line!=Data_base().films_sheet[f'J2'].value+2:
            if Data_base().films_sheet[f'A{line}'].value == id: return line
            line+=1

    def current_film_icon(self, id):
        icon = CTk.CTkImage(light_image=Image.open(f'img/films/{id}.png'),
                                  dark_image=Image.open(f'img/films/{id}.png'),
                                  size=(256, 256))
        return icon

    def current_film_name(self, id):
        current_line = self.current_line(self, id)
        name = Data_base().films_sheet[f'B{current_line}'].value
        return name

    def current_film_description(self, id):
        current_line = self.current_line(self, id)
        description = Data_base().films_sheet[f'C{current_line}'].value
        return description

    def current_film_duration(self, id):
        current_line = self.current_line(self, id)
        duration = Data_base().films_sheet[f'E{current_line}'].value
        return duration

    def current_film_age_limit(self, id):
        current_line = self.current_line(self, id)
        age_limit = f'{Data_base().films_sheet[f"F{current_line}"].value}+'
        return age_limit
        
        
if __name__ == '__main__':
    try: 
        Data_base()
    except: 
        messagebox.showerror('Ошибка', 'Ошибка загрузки базы данных')
        exit()
    try: 
        User()
    except: 
        messagebox.showerror('Ошибка', 'Ошибка загрузки данных пользователя')
        exit()
    try:
        Login_window()
    except: 
        messagebox.showerror('Ошибка', 'Ошибка загрузки интерфейса Tkinter')
        exit()